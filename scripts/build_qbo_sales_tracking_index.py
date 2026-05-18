#!/usr/bin/env python3
"""Build a derived invoice-number index from Rowboat sales tracking XLSX files.

Reads Joseph-provided sales tracking spreadsheets and writes a derived JSON index
outside original evidence. Original XLSX files are never modified.
"""
from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

LOCKER_ROOT = Path("/Volumes/iseepatterns-evidence/ISEEPATTERNS_LOCKER")
OUTPUT_PATH = LOCKER_ROOT / "analysis/uploads/qbo_unallocated_invoice_uploads/sales_tracking_invoice_index.json"
SOURCE_FILES = [
    LOCKER_ROOT / "lawmodel1/data/FINANCIAL_LOCKER/ROWBOAT_CREATIVE_SALES_LOCKER/sales/05-28 open jobs_1cCvzR5P8GmlAbzYLpvLpiQqDeniBjZ3wYRzc1L6Y21s.xlsx",
    LOCKER_ROOT / "lawmodel1/data/FINANCIAL_LOCKER/ROWBOAT_CREATIVE_SALES_LOCKER/sales/Sales Figures 18 19_16-GCCL0RYtVrtRW0D9G41hBSnEzK4e4gaf-fa_SASlA.xlsx",
    LOCKER_ROOT / "lawmodel1/data/FINANCIAL_LOCKER/ROWBOAT_CREATIVE_SALES_LOCKER/sales/Weekly Sales Goals 2019_1YmljLrfTXb6wv4LH8NgILC7BgJk-UGZyhkT32Q2cOLw.xlsx",
]

HEADER_ALIASES = {
    "invoice": {"num", "inv", "order #", "order no", "order number"},
    "customer": {"name", "company", "customer", "client"},
    "salesperson": {"sales rep", "sales person", "salesperson"},
    "date": {"date", "date ordered"},
    "due_date": {"due date"},
    "amount": {"amount", "total", "amount billed"},
    "open_balance": {"open balance"},
    "ar_paid": {"a/r paid", "ar paid"},
    "customer_po": {"customer po"},
    "memo": {"memo/description", "customer/vendor message", "job name", "notes"},
}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def serialize(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def invoice_key(value: Any) -> str:
    text = serialize(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    text = text.replace(",", "").strip()
    return text


def find_header_row(sheet) -> tuple[int | None, list[str]]:
    best_row = None
    best_headers: list[str] = []
    best_score = 0
    max_row = sheet.max_row or 0
    for row_index in range(1, min(max_row, 30) + 1):
        values = [norm_header(cell.value) for cell in sheet[row_index]]
        score = 0
        for aliases in HEADER_ALIASES.values():
            if any(value in aliases for value in values):
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_index
            best_headers = values
    if best_score >= 2:
        return best_row, best_headers
    return None, []


def alias_positions(headers: list[str]) -> dict[str, int]:
    positions: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(headers):
            if header in aliases:
                positions[field] = idx
                break
    return positions


def row_has_data(values: list[Any]) -> bool:
    return any(serialize(value) for value in values)


def main() -> None:
    by_invoice: dict[str, list[dict[str, str]]] = {}
    sources: list[dict[str, Any]] = []
    total_hits = 0

    for source in SOURCE_FILES:
        wb = load_workbook(source, data_only=True, read_only=False)
        source_entry: dict[str, Any] = {
            "path": str(source),
            "sha256": sha256_file(source),
            "sheets": [],
        }
        for sheet in wb.worksheets:
            header_row, headers = find_header_row(sheet)
            if header_row is None:
                source_entry["sheets"].append({"sheet": sheet.title, "header_row": None, "indexed_rows": 0, "note": "no supported header detected"})
                continue
            positions = alias_positions(headers)
            indexed_rows = 0
            if "invoice" not in positions:
                source_entry["sheets"].append({"sheet": sheet.title, "header_row": header_row, "indexed_rows": 0, "note": "no invoice/order column detected", "headers": headers})
                continue
            for excel_row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                values = list(row)
                if not row_has_data(values):
                    continue
                key = invoice_key(values[positions["invoice"]] if positions["invoice"] < len(values) else "")
                if not key:
                    continue
                hit = {
                    "source_path": str(source),
                    "source_sha256": source_entry["sha256"],
                    "sheet": sheet.title,
                    "excel_row_number": str(excel_row_number),
                    "invoice_or_order_number": key,
                    "customer_name": serialize(values[positions["customer"]]) if "customer" in positions and positions["customer"] < len(values) else "",
                    "salesperson_raw": serialize(values[positions["salesperson"]]) if "salesperson" in positions and positions["salesperson"] < len(values) else "",
                    "date": serialize(values[positions["date"]]) if "date" in positions and positions["date"] < len(values) else "",
                    "due_date": serialize(values[positions["due_date"]]) if "due_date" in positions and positions["due_date"] < len(values) else "",
                    "amount": serialize(values[positions["amount"]]) if "amount" in positions and positions["amount"] < len(values) else "",
                    "open_balance": serialize(values[positions["open_balance"]]) if "open_balance" in positions and positions["open_balance"] < len(values) else "",
                    "ar_paid": serialize(values[positions["ar_paid"]]) if "ar_paid" in positions and positions["ar_paid"] < len(values) else "",
                    "customer_po": serialize(values[positions["customer_po"]]) if "customer_po" in positions and positions["customer_po"] < len(values) else "",
                    "memo": serialize(values[positions["memo"]]) if "memo" in positions and positions["memo"] < len(values) else "",
                    "basis": "matched QBO doc_number to sales tracking invoice/order field",
                }
                by_invoice.setdefault(key, []).append(hit)
                indexed_rows += 1
                total_hits += 1
            source_entry["sheets"].append({
                "sheet": sheet.title,
                "header_row": header_row,
                "indexed_rows": indexed_rows,
                "headers": [h for h in headers if h],
            })
        sources.append(source_entry)
        wb.close()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema_version": 1,
        "purpose": "Derived invoice-number index from Joseph-identified Rowboat sales tracking XLSX files for QBO unallocated review UI suggestions.",
        "created_at_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "storage_note": "Derived analysis artifact; original XLSX evidence files are not modified.",
        "controlling_salesperson_field": "QBO Invoice CustomField DefinitionId=2 remains controlling; sales tracking hits are source-referenced suggestions only.",
        "output_path": str(OUTPUT_PATH),
        "source_files": sources,
        "total_unique_invoice_keys": len(by_invoice),
        "total_indexed_hits": total_hits,
        "by_invoice": by_invoice,
    }
    text = json.dumps(payload, indent=2, sort_keys=True)
    OUTPUT_PATH.write_text(text, encoding="utf-8")
    print(json.dumps({
        "output_path": str(OUTPUT_PATH),
        "sha256": hashlib.sha256(text.encode("utf-8")).hexdigest(),
        "total_unique_invoice_keys": len(by_invoice),
        "total_indexed_hits": total_hits,
    }, indent=2))


if __name__ == "__main__":
    main()
